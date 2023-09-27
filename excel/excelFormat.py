import pandas as pd
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from PIL import Image
from io import BytesIO
import requests


def excelFormat(excel_path):
    
    wb = load_workbook(filename=excel_path)
    ws = wb.active    
    
    # 读取Pandas DataFrame
    #df = pd.read_excel(excel_path)
    
    #column_order = ["分类1", "分类2", "分类3", "图", "标题", "描述", "价格", "rating", "人数", "地址",'男士评分','女士评分','processed','translated']
    #final_df = df[column_order]
    
    # 将最终DataFrame保存到Excel
    #with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
    #    writer.book = wb
    #   final_df.to_excel(writer, index=False)
    
    # 设置每行的宽度
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 20
    ws.column_dimensions['F'].width = 40
    ws.column_dimensions['J'].width = 30

    wb.save(excel_path)
    
    processed = 0
    
    for row in ws.iter_rows(min_col=4, max_col=4, min_row=2):  # 假设“图”列在第一列
        for cell in row:
            image_url = cell.value  # 从单元格获取图片URL
            # 使用requests库从网络下载图片
            response = requests.get(image_url)
            if response.status_code == 200:
                img_data = BytesIO(response.content)

                # 使用PIL库来打开和调整图片
                with Image.open(img_data) as img:
                    w, h = img.size
                    scaling_factor = min(200/w, 200/h)
                    img = img.resize((int(w * scaling_factor), int(h * scaling_factor)))

                    img_byte_arr = BytesIO()
                    img.save(img_byte_arr, format='JPEG')
                    img_byte_arr = img_byte_arr.getvalue()

                # 将图片插入到Excel单元格
                img = openpyxl.drawing.image.Image(BytesIO(img_byte_arr))
                ws.add_image(img, cell.coordinate)
                
                # 调整单元格大小
                ws.row_dimensions[cell.row].height = 8 * 20  # 约等于200像素
                ws.column_dimensions[openpyxl.utils.get_column_letter(cell.column)].width = 25  # 约等于200像素
                
                processed += 1
                print("Processed: ", processed)

    for row in ws.iter_rows(min_row=2, min_col= 10, max_col=10):  # 假设第一行是表头，从第二行开始
        for cell in row:
            if isinstance(cell.value, str) and 'http' in cell.value:  # 检查单元格是否包含'http'
                cell.hyperlink = cell.value  # 将单元格文本设置为超链接

    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(horizontal='left', vertical='center', wrapText=True)

    # 保存修改后的Excel文件
    wb.save(excel_path)