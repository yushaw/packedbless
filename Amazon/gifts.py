from concurrent.futures import ThreadPoolExecutor
import pandas as pd
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
import os
import pickle
from selenium import webdriver
from selenium.webdriver.common.by import By
from bs4 import BeautifulSoup
import time
import random
import captcha.amazon as captcha


def get_details(product_link): 
    
    driver = webdriver.Chrome() 
    
     # 随机延迟
    time.sleep(random.uniform(2.0, 6.0))
    
    driver.get(product_link)  
    
    # Check for captcha
    if "captcha" in driver.page_source.lower():
        captcha_solved = captcha.solve_captcha(driver)
        if not captcha_solved:
            print("Failed to solve captcha. Product link: ", product_link)
            driver.quit()
            return  {"分类1": "NaN",
                    "分类2": "NaN",
                    "分类3": "NaN",
                    "描述": "NaN",
                    "rating": "NaN"}
  
    # 提取详细信息
    try:
        detail_soup = BeautifulSoup(driver.page_source, 'html.parser')
    except Exception as e:
        print("Error occurred while parsing page source: ", e)
        return None
    
    categories = detail_soup.select("#wayfinding-breadcrumbs_container a")
    
    try:
        cat1, cat2, cat3 = categories[0].text.strip(), categories[1].text.strip(), categories[2].text.strip()
    except:
        print("Problem with categories. Product link: ", product_link)
        cat1, cat2, cat3 = None, None, None
        pass
        
    
    try:
        description_items = detail_soup.select("#feature-bullets li")
    except Exception as e:
        print("Error occurred while selecting description items: ", e)
        description_items = None
        pass
    
    description = " ".join(item.text.strip() for item in description_items)
    
    parent_div = detail_soup.find('div', {'id': 'averageCustomerReviews'})
    try:
        rating_element = parent_div.find('span', {'class': 'a-size-base a-color-base'})
    except:
        try:
            rating_element = parent_div.find('span', {'class': 'a-icon-alt'})
        except:
            print("Problem with detailed rating. Product link: ", product_link)
            rating_element = None
            pass
                        
    if rating_element:
        rating = rating_element.text.strip()
        try:
            # 提取数字并转换为 int 类型
            rating = float(rating)
        except ValueError:
            # 如果转换失败，设置为 None
            print("Problem with detailed rating. Product link: ", product_link)
            rating = None
            pass
    else:
        rating = "N/A"
    
    # 随机延迟
    # time.sleep(random.uniform(2.0, 6.0))    
    driver.quit()
    return {"分类1": cat1,
                    "分类2": cat2,
                    "分类3": cat3,
                    "描述": description,
                    "rating": rating}
    

def phaseOne(main_link,excel_path, toCollect):
    
    # 创建目录
    directory = os.path.dirname(excel_path)
    if not os.path.exists(directory):
        os.makedirs(directory)
    
    # Initialize an empty list to store the data
    data_list = []

    # 初始化Selenium Webdriver
    driver = webdriver.Chrome()

    # 访问网站
    driver.get(main_link)

    # 用于去重的集合
    product_links = set()

    # 爬取成功的商品数量
    success_count = 0

    # 上限数量
    limit = toCollect
    
    previous_elements_count = 0

    # 模拟无限滚动
    while success_count < limit:
        soup = BeautifulSoup(driver.page_source, 'html.parser')
        
        for section in soup.select("[id^='gcx-gf-section-']"):
            for product in section.select(".d4xojt-0"):
                if success_count >= limit:
                    break
                
                unique_id = product.get('class')[-1]
                
                # 在新标签页中打开商品详细页面
                product_link = "https://www.amazon.com" + product.select_one("a")["href"]

                # 去重
                if product_link in product_links:
                    continue
                product_links.add(product_link)
                
                if success_count >= limit:
                    break
                
                # 提取信息
                title = product.select_one("[data-test='product'] span").text
                
                try:
                    price = product.select_one("[data-test='price']").text
                except:
                    price = None
                
                try:
                    # 提取数字并转换为 float 类型
                    price = float(price.replace("$", ""))
                except:
                    # 如果转换失败，设置为 None
                    print("Problem with price. Product link: ", product_link)
                    price = None
                    pass

                reviews = product.select_one(".ptijou-1").text  # 根据实际情况调整
                reviews = reviews.replace(',', '')
                try:
                    # 提取数字并转换为 int 类型
                    reviews = int(reviews)
                except:
                    # 如果转换失败，设置为 None
                    print("Problem with reviews. Product link: ", product_link)
                    reviews = None
                    raise ValueError(str(e))
                
                rating_tag = product.find('i', {'class': lambda x: x and 'a-star-small-' in x})
                # 提取评分
                if rating_tag:
                    for class_name in rating_tag['class']:
                        if 'a-star-small-' in class_name:
                            raw_rating = class_name.replace('a-star-small-', '').replace('-', '.')
                            raw_rating = float(raw_rating)
                            break
                else:
                    raw_rating = None

                img_tag = product.find('img', {'data-a-hires': True})
                if img_tag:
                    image_url = img_tag['data-a-hires']
                else:
                    print("Image URL with attribute 'data-a-hires' not found.")
                    
                data_list.append({
                    "图": image_url,  # This will be the image URL, actual image insertion happens later
                    "标题": title,
                    "价格": price,
                    "人数": reviews,
                    "地址": product_link
                })
                                
                # 更新成功数量
                success_count += 1
                
                print(f"成功爬取 {success_count} / {toCollect}个商品")
                
                # 检查是否达到上限
                if success_count >= limit:
                    break
            
            if success_count >= limit:
                break
        
        # 随机延迟
        time.sleep(random.uniform(1.0, 3.0))
        
        # 滚动页面
        driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
        
        # 随机延迟，等待页面加载
        time.sleep(random.uniform(3.0, 5.0))

        # 计算当前页面元素数量（以您要抓取的商品为例）
        current_elements_count = len(driver.find_elements(By.CSS_SELECTOR, "[id^='gcx-gf-section-']"))

        # 如果数量没有变化，说明没有更多内容可以加载
        if current_elements_count == previous_elements_count:
            print("Reached the end of content.")
            break
        else:
            previous_elements_count = current_elements_count
    
    
    df = pd.DataFrame(data_list)

    # Save the DataFrame to an Excel file
    try:
        df.to_excel(excel_path, index=False)
    except ValueError as e:
        print("Error while saving to Excel file:", str(e))
        raise ValueError(str(e))


def phaseTwo(excel_path, max_workers):
        
    wb = load_workbook(filename=excel_path)
    
    if len(wb.sheetnames) > 0:
        wb.active = 0
    else:
        raise ValueError("Workbook has no sheets.")

    ws = wb.active    
    
    # 初始化Selenium Webdriver
    driver = webdriver.Chrome()
    
    # 读取Pandas DataFrame
    df = pd.read_excel(excel_path)
        
    # 检查是否有临时数据
    temp_dir = './temp'
    details_file = os.path.join(temp_dir, 'details.pkl')
    completed_tasks_file = os.path.join(temp_dir, 'completed_tasks.pkl')
    
    if not os.path.exists(temp_dir):
        os.mkdir(temp_dir)

    try:
        with open(details_file, 'rb') as f:
            details = pickle.load(f)
        with open(completed_tasks_file, 'rb') as f:
            completed_tasks = pickle.load(f)
    except FileNotFoundError:
        details = []
        completed_tasks = 0
    
    total_tasks = len(df["地址"])
    save_interval = total_tasks // 9  # 计算保存间隔

    # 执行任务
    with ThreadPoolExecutor(max_workers) as executor:
        for result in executor.map(get_details, df["地址"][completed_tasks:]):
            details.append(result)
            completed_tasks += 1

            # 根据保存间隔进行保存
            if completed_tasks % save_interval == 0:
                with open(details_file, 'wb') as f:
                    pickle.dump(details, f)
                with open(completed_tasks_file, 'wb') as f:
                    pickle.dump(completed_tasks, f)

            print(f"Processed: {completed_tasks}")

    details_df = pd.DataFrame(details)
    final_df = pd.concat([df, details_df], axis=1)
    
    print("final_df head:", final_df.head())

    # Check data types and structure
    print("df info:")
    final_df.info()

    column_order = ["分类1", "分类2", "分类3", "图", "标题", "描述", "价格", "rating", "人数", "地址"]
    final_df = final_df[column_order]
    print("final_df head:", final_df.head())

    # Check data types and structure
    print("df info:")
    final_df.info()
    
    # 创建一个新的Excel工作簿
    wb = Workbook()
    ws = wb.active  # 获取活动工作表

    # 假设final_df是您要保存的DataFrame
    for col_num, column_title in enumerate(final_df.columns, 1):
        col_letter = get_column_letter(col_num)
        ws[f"{col_letter}1"] = column_title

    # 写入数据
    for row_num, row_data in enumerate(final_df.values, 2):
        for col_num, cell_value in enumerate(row_data, 1):
            col_letter = get_column_letter(col_num)
            ws[f"{col_letter}{row_num}"] = cell_value

    # 保存工作簿
    wb.save(excel_path)

    
    # 任务完成，删除临时数据
    os.remove(details_file)
    os.remove(completed_tasks_file)
    
    driver.quit()
    
    
def get_gifts(main_link, excel_path, toCollect=5, max_workers=3):
    phaseOne(main_link, excel_path, toCollect)
    phaseTwo(excel_path, max_workers)