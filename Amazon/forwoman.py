# Import required libraries for Excel operations
import pandas as pd
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from PIL import Image
from io import BytesIO
import base64
import requests


def forWoman(toCollect=1000):
    # Initialize an empty list to store the data
    data_list = []

    from selenium import webdriver
    from bs4 import BeautifulSoup
    import time
    import random

    # 初始化Selenium Webdriver
    driver = webdriver.Chrome()

    # 访问网站
    driver.get("https://www.amazon.com/gcx/Gifts-for-Women/gfhz/category?categoryId=adult-female")

    # 用于去重的集合
    product_links = set()

    # 爬取成功的商品数量
    success_count = 0

    # 上限数量
    limit = toCollect

    # 模拟无限滚动
    while success_count < limit:
        soup = BeautifulSoup(driver.page_source, 'html.parser')
        
        for section in soup.select("[id^='gcx-gf-section-']"):
            for product in section.select(".d4xojt-0"):
                unique_id = product.get('class')[-1]
                
                # 在新标签页中打开商品详细页面
                product_link = "https://www.amazon.com" + product.select_one("a")["href"]
                print("current product", product_link)

                # 去重
                if product_link in product_links:
                    continue
                product_links.add(product_link)
                
                # 提取信息
                title = product.select_one("[data-test='product'] span").text
                price = product.select_one("[data-test='price']").text
                try:
                    # 提取数字并转换为 float 类型
                    price = float(price.replace("$", ""))
                except ValueError:
                    # 如果转换失败，设置为 None
                    pass

                reviews = product.select_one(".ptijou-1").text  # 根据实际情况调整
                reviews = reviews.replace(',', '')
                try:
                    # 提取数字并转换为 int 类型
                    reviews = int(reviews)
                except ValueError:
                    # 如果转换失败，设置为 None
                    reviews = None
                
                
                rating_tag = product.find('i', {'class': lambda x: x and 'a-star-small-' in x})
                # 提取评分
                if rating_tag:
                    for class_name in rating_tag['class']:
                        if 'a-star-small-' in class_name:
                            raw_rating = class_name.replace('a-star-small-', '').replace('-', '.')
                            raw_rating = float(raw_rating)
                            break
                else:
                    raw_rating = None

                img_tag = product.find('img', {'data-a-hires': True})
                if img_tag:
                    image_url = img_tag['data-a-hires']
                else:
                    print("Image URL with attribute 'data-a-hires' not found.")
                    
                driver.execute_script(f"window.open('{product_link}', '_blank');")
                
                # 切换到新标签页
                driver.switch_to.window(driver.window_handles[-1])
                
                # 提取详细信息
                detail_soup = BeautifulSoup(driver.page_source, 'html.parser')
                categories = detail_soup.select("#wayfinding-breadcrumbs_container a")
                
                try:
                    cat1, cat2, cat3 = categories[0].text.strip(), categories[1].text.strip(), categories[2].text.strip()
                except:
                    cat1, cat2, cat3 = None, None, None

                description_items = detail_soup.select("#feature-bullets li")
                description = " ".join(item.text.strip() for item in description_items)
                
                parent_div = detail_soup.find('div', {'id': 'averageCustomerReviews'})
                
                try:
                    rating_element = parent_div.find('span', {'class': 'a-size-base a-color-base'})
                except:
                    try:
                        rating_element = parent_div.find('span', {'class': 'a-icon-alt'})
                    except:
                        pass
                
                if rating_element:
                    rating = rating_element.text.strip()
                    try:
                        # 提取数字并转换为 int 类型
                        rating = float(rating)
                    except:
                        # 如果转换失败，设置为 None
                        pass
                else:
                    rating = "N/A"

                # print(f"Title: {title}, Price: {price}, Rating: {rating}, Reviews: {reviews}")
                # print(f"Categories: {cat1} > {cat2} > {cat3}")
                # print(f"Description: {description}")
                
                # Inside your loop where you collect data, instead of printing, append it to the data_list
                # For example:
                data_list.append({
                    "分类1": cat1,
                    "分类2": cat2,
                    "分类3": cat3,
                    "图": image_url,  # This will be the image URL, actual image insertion happens later
                    "标题": title,
                    "描述": description,
                    "价格": price,
                    "rating": rating,
                    "人数": reviews,
                    "地址": product_link
                })
                
                # 关闭当前标签页并返回到原始标签页
                driver.close()
                driver.switch_to.window(driver.window_handles[0])
                
                # 更新成功数量
                success_count += 1
                
                # 检查是否达到上限
                if success_count >= limit:
                    break
        
        # 随机延迟
        time.sleep(random.uniform(1.0, 3.0))
        
        # 滚动页面
        driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")



    # Create a DataFrame from the collected data
    df = pd.DataFrame(data_list)

    # Save the DataFrame to an Excel file
    excel_path = './data/amazon_woman.xlsx'
    df.to_excel(excel_path, index=False)

    # 加载之前保存的Excel文件
    wb = load_workbook(excel_path)

    # 选择要操作的工作表
    ws = wb.active

    # 设置每行的宽度
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 20
    ws.column_dimensions['F'].width = 40
    ws.column_dimensions['J'].width = 30

    for row in ws.iter_rows(min_col=4, max_col=4, min_row=2):  # 假设“图”列在第一列
        for cell in row:
            image_url = cell.value  # 从单元格获取图片URL
            # 使用requests库从网络下载图片
            response = requests.get(image_url)
            if response.status_code == 200:
                img_data = BytesIO(response.content)

                # 使用PIL库来打开和调整图片
                with Image.open(img_data) as img:
                    w, h = img.size
                    scaling_factor = min(200/w, 200/h)
                    img = img.resize((int(w * scaling_factor), int(h * scaling_factor)))

                    img_byte_arr = BytesIO()
                    img.save(img_byte_arr, format='JPEG')
                    img_byte_arr = img_byte_arr.getvalue()

                # 将图片插入到Excel单元格
                img = openpyxl.drawing.image.Image(BytesIO(img_byte_arr))
                ws.add_image(img, cell.coordinate)
                
                # 调整单元格大小
                ws.row_dimensions[cell.row].height = 8 * 20  # 约等于200像素
                ws.column_dimensions[openpyxl.utils.get_column_letter(cell.column)].width = 25  # 约等于200像素

    for row in ws.iter_rows(min_row=2, min_col= 10, max_col=10):  # 假设第一行是表头，从第二行开始
        for cell in row:
            if isinstance(cell.value, str) and 'http' in cell.value:  # 检查单元格是否包含'http'
                cell.hyperlink = cell.value  # 将单元格文本设置为超链接

    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(horizontal='left', vertical='center', wrapText=True)

    # 保存修改后的Excel文件
    wb.save(excel_path)
